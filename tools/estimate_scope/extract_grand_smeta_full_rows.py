#!/usr/bin/env python3
"""Extract full row composition from Grand-Smeta/estimate XLSX exports.

This tool is intentionally different from extract_estimate_scope.py:
- extract_estimate_scope.py builds a compact register of likely scope rows;
- this script preserves every meaningful row, including resources, subtotals,
  overhead/profit rows and "total by position" rows.

Use it when the commercial question is not only "what works are listed", but
"what rows form the estimate price and which parent position they belong to".
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

POSITION_RE = re.compile(r"^\s*\d+\s*$")
TOTAL_ESTIMATE_RE = re.compile(r"всего\s+по\s+смете", re.IGNORECASE)
TOTAL_POSITION_RE = re.compile(r"всего\s+по\s+позиции", re.IGNORECASE)
DIRECT_COST_RE = re.compile(r"итого\s+прямые\s+затраты", re.IGNORECASE)
OVERHEAD_RE = re.compile(r"\b(нр|сп)\b|накладн|сметн\s+приб", re.IGNORECASE)
LABOR_RE = re.compile(r"труд|фот|заработ", re.IGNORECASE)
RESOURCE_GROUP_RE = re.compile(r"материальн|машин|механизм|оборудован|ресурс", re.IGNORECASE)

DEFAULT_COLUMNS = [
    "source_file",
    "sheet",
    "row_number",
    "parent_position",
    "position",
    "basis",
    "name",
    "unit",
    "quantity_per_unit",
    "quantity_total",
    "base_price",
    "coefficient",
    "current_price",
    "row_cost_or_total",
    "row_type",
    "classification_note",
    "row_text",
]


@dataclass
class SourceSummary:
    source_file: str
    sheet: str
    non_empty_rows: int = 0
    parent_positions: int = 0
    estimate_total: object | None = None


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def text_for_row(values: Sequence[object]) -> str:
    return " | ".join(normalize(value) for value in values if normalize(value))


def is_number_like(value: object) -> bool:
    if isinstance(value, (int, float)):
        return True
    text = normalize(value).replace(" ", "").replace(",", ".")
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def first_non_empty(values: Sequence[object], start: int = 0, end: int | None = None) -> object:
    limit = len(values) if end is None else min(end, len(values))
    for index in range(start, limit):
        if normalize(values[index]):
            return values[index]
    return ""


def detect_column(values: Sequence[object], candidates: Sequence[int]) -> object:
    for one_based in candidates:
        index = one_based - 1
        if 0 <= index < len(values) and normalize(values[index]):
            return values[index]
    return ""


def detect_position(values: Sequence[object]) -> str:
    # Most estimate exports put the row number/position in the first columns.
    for value in values[:3]:
        text = normalize(value)
        if POSITION_RE.match(text):
            return text
    return ""


def detect_basis(values: Sequence[object]) -> object:
    # Grand-Smeta XLSX exports vary, but basis is usually near columns B-D.
    return detect_column(values, [2, 3, 4])


def detect_name(values: Sequence[object]) -> object:
    # Prefer columns where Grand-Smeta usually stores work/resource names.
    for one_based in [4, 5, 6, 7, 3, 2]:
        index = one_based - 1
        if 0 <= index < len(values):
            text = normalize(values[index])
            if text and not POSITION_RE.match(text):
                return values[index]
    return first_non_empty(values)


def detect_unit(values: Sequence[object]) -> object:
    # Typical unit columns are around G-I in many Russian estimate exports.
    return detect_column(values, [7, 8, 9, 6])


def numeric_tail(values: Sequence[object]) -> list[object]:
    return [value for value in values if is_number_like(value)]


def detect_quantities_and_prices(values: Sequence[object]) -> tuple[object, object, object, object, object, object]:
    numbers = numeric_tail(values)
    quantity_per_unit = numbers[0] if len(numbers) > 0 else ""
    quantity_total = numbers[1] if len(numbers) > 1 else ""
    base_price = numbers[-4] if len(numbers) >= 4 else ""
    coefficient = numbers[-3] if len(numbers) >= 3 else ""
    current_price = numbers[-2] if len(numbers) >= 2 else ""
    row_cost_or_total = numbers[-1] if numbers else ""
    return quantity_per_unit, quantity_total, base_price, coefficient, current_price, row_cost_or_total


def classify_row(values: Sequence[object], parent_position: str, position: str) -> tuple[str, str]:
    row_text = text_for_row(values)
    lower = row_text.lower()

    if TOTAL_ESTIMATE_RE.search(row_text):
        return "Итог по смете", "Общий итог сметы; не является отдельным видом работ"
    if TOTAL_POSITION_RE.search(row_text):
        return "Итого по позиции", "Итог стоимости родительской сметной позиции"
    if DIRECT_COST_RE.search(row_text):
        return "Итого прямые затраты", "Промежуточный итог в составе позиции"
    if OVERHEAD_RE.search(row_text):
        return "НР / СП", "Накладные расходы или сметная прибыль в составе позиции"
    if LABOR_RE.search(row_text):
        return "Труд / ФОТ", "Строка трудозатрат или фонда оплаты труда"
    if position:
        if parent_position == position:
            return "Основная сметная позиция", "Сравнивать с ВОР/РД на этом уровне"
        return "Ресурсная строка с номером", "Ресурс или вложенная строка внутри родительской позиции"
    if RESOURCE_GROUP_RE.search(lower):
        return "Группа ресурсов", "Заголовок ресурсного блока"
    if parent_position:
        return "Ресурс / вложенная строка", "Не сравнивать с ВОР как самостоятельную работу без родительской позиции"
    return "Прочая строка", "Служебная или нераспознанная строка"


def iter_xlsx_files(input_path: Path) -> Iterable[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() == ".xlsx":
            yield input_path
        return
    for path in sorted(input_path.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        yield path


def extract_workbook(path: Path) -> tuple[list[dict[str, object]], list[SourceSummary]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, object]] = []
    summaries: list[SourceSummary] = []

    for sheet in workbook.worksheets:
        parent_position = ""
        summary = SourceSummary(source_file=str(path), sheet=sheet.title)

        for row_number, excel_row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(excel_row)
            if not any(normalize(value) for value in values):
                continue

            summary.non_empty_rows += 1
            position = detect_position(values)
            if position:
                parent_position = position
                summary.parent_positions += 1

            row_type, note = classify_row(values, parent_position, position)
            q1, q2, p1, coef, p2, total = detect_quantities_and_prices(values)
            row_text = text_for_row(values)

            if TOTAL_ESTIMATE_RE.search(row_text):
                summary.estimate_total = total

            rows.append(
                {
                    "source_file": str(path),
                    "sheet": sheet.title,
                    "row_number": row_number,
                    "parent_position": parent_position,
                    "position": position,
                    "basis": detect_basis(values),
                    "name": detect_name(values),
                    "unit": detect_unit(values),
                    "quantity_per_unit": q1,
                    "quantity_total": q2,
                    "base_price": p1,
                    "coefficient": coef,
                    "current_price": p2,
                    "row_cost_or_total": total,
                    "row_type": row_type,
                    "classification_note": note,
                    "row_text": row_text,
                }
            )

        summaries.append(summary)

    workbook.close()
    return rows, summaries


def write_xlsx(output_path: Path, rows: list[dict[str, object]], summaries: list[SourceSummary]) -> None:
    workbook = Workbook()
    rows_sheet = workbook.active
    rows_sheet.title = "FullRows"
    rows_sheet.append(DEFAULT_COLUMNS)
    for row in rows:
        rows_sheet.append([row.get(column, "") for column in DEFAULT_COLUMNS])

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["source_file", "sheet", "non_empty_rows", "parent_positions", "estimate_total"])
    for summary in summaries:
        summary_sheet.append(
            [summary.source_file, summary.sheet, summary.non_empty_rows, summary.parent_positions, summary.estimate_total]
        )

    method_sheet = workbook.create_sheet("Method")
    method_sheet.append(["Правило", "Описание"])
    method_sheet.append([
        "Уровень сравнения",
        "С ВОР и РД сравниваются родительские сметные позиции; ресурсные строки не считаются самостоятельными работами.",
    ])
    method_sheet.append([
        "Назначение",
        "Лист FullRows сохраняет полный состав строк сметы: работы, ресурсы, НР/СП, итоги по позиции и общий итог.",
    ])
    method_sheet.append([
        "Кодировка",
        "Для CSV используйте UTF-8-SIG, для XLSX кириллица хранится как Unicode.",
    ])

    workbook.save(output_path)


def write_csv(output_path: Path, rows: list[dict[str, object]]) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=DEFAULT_COLUMNS, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="XLSX file or folder with XLSX estimate exports")
    parser.add_argument("--output", required=True, help="Output .xlsx or .csv file")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"Input does not exist: {input_path}", file=sys.stderr)
        return 2

    all_rows: list[dict[str, object]] = []
    all_summaries: list[SourceSummary] = []
    for xlsx_path in iter_xlsx_files(input_path):
        rows, summaries = extract_workbook(xlsx_path)
        all_rows.extend(rows)
        all_summaries.extend(summaries)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        write_csv(output_path, all_rows)
    else:
        write_xlsx(output_path, all_rows, all_summaries)

    print(f"Extracted {len(all_rows)} rows from {len(all_summaries)} sheets -> {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
