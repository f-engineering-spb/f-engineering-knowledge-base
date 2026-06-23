#!/usr/bin/env python3
"""Extract compact estimate/VOR scope rows from XLSX workbooks.

The script is intentionally generic and does not contain project-specific data.
It scans workbooks, keeps rows that match keywords, classifies likely facade
scope categories, and writes a traceable register for further review.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover - exercised by user environment
    raise SystemExit(
        "Missing dependency: openpyxl. Install requirements.txt or use the "
        "Codex workspace Python runtime with bundled spreadsheet packages."
    ) from exc


DEFAULT_KEYWORDS = [
    "вентилируем",
    "вентфасад",
    "фасад",
    "керамогранит",
    "гранит",
    "камень",
    "цокол",
    "откос",
    "обрамлен",
    "обрамлени",
    "отлив",
    "оцинк",
    "листов",
    "парапет",
    "покрыт",
]


UNIT_HINTS = {
    "м2",
    "м²",
    "м",
    "п.м",
    "пог.м",
    "шт",
    "т",
    "кг",
    "компл",
}


@dataclass
class ScopeRow:
    source_file: str
    sheet: str
    row_number: int
    corpus: str
    category: str
    position: str
    unit: str
    quantity: float | None
    formula_or_note: str
    row_text: str


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not text:
        return None
    if re.fullmatch(r"[-+]?\d+(\.\d+)?", text):
        return float(text)
    return None


def detect_unit(cells: Sequence[object]) -> str:
    for cell in cells:
        text = normalize(cell).lower()
        if text in UNIT_HINTS:
            return text
    return ""


def detect_quantity(cells: Sequence[object], unit_index: int | None) -> float | None:
    if unit_index is not None:
        for cell in cells[unit_index + 1 : unit_index + 5]:
            number = parse_number(cell)
            if number is not None:
                return number
    numbers = [parse_number(cell) for cell in cells]
    numbers = [number for number in numbers if number is not None]
    return numbers[-1] if numbers else None


def classify(text: str) -> str:
    lower = text.lower()
    if "демонтаж" in lower or "разборк" in lower:
        return "Demolition"
    has_granite = "гранит" in lower and "керамогранит" not in lower
    if has_granite or "камень" in lower or "цокол" in lower:
        return "Plinth / granite / stone"
    if "откос" in lower or "обрамлен" in lower or "обрамлени" in lower:
        return "Opening slopes / framing"
    if "отлив" in lower:
        return "Window flashings"
    if "оцинк" in lower or "листов" in lower or "мелкие покрыт" in lower:
        return "Galvanized sheet metal / minor coverings"
    if "парапет" in lower or "покрытие парапет" in lower:
        return "Parapets / copings"
    if "вентилируем" in lower or "вентфасад" in lower or "керамогранит" in lower:
        return "NVF / porcelain stoneware"
    if "фасад" in lower:
        return "Other facade-related"
    return "Other"


def detect_corpus(path: Path, sheet: str, row_text: str) -> str:
    text = f"{path.name} {sheet} {row_text}"
    match = re.search(r"(АР|AR)\s*([123])", text, re.IGNORECASE)
    if match:
        return f"АР{match.group(2)}"
    match = re.search(r"\b(02-01|02-02|03-01)\b", text)
    if match:
        return {"02-01": "АР1", "02-02": "АР2", "03-01": "АР3"}[match.group(1)]
    return ""


def best_position(cells: Sequence[object]) -> str:
    texts = [normalize(cell) for cell in cells if normalize(cell)]
    candidates = [text for text in texts if len(text) > 12 and not re.fullmatch(r"[\d., ()+\-*/]+", text)]
    if candidates:
        return max(candidates, key=len)
    return " | ".join(texts[:4])


def row_matches(row_text: str, keywords: Sequence[str]) -> bool:
    lower = row_text.lower()
    return any(keyword.lower() in lower for keyword in keywords)


def iter_xlsx_files(input_path: Path) -> Iterable[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".xlsx":
        yield input_path
        return
    for path in sorted(input_path.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        yield path


def extract_file(path: Path, keywords: Sequence[str]) -> list[ScopeRow]:
    rows: list[ScopeRow] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        for row_number, cells in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(cells)
            row_text = " | ".join(normalize(cell) for cell in values if normalize(cell))
            if not row_text or not row_matches(row_text, keywords):
                continue
            unit_index = None
            for index, cell in enumerate(values):
                if normalize(cell).lower() in UNIT_HINTS:
                    unit_index = index
                    break
            rows.append(
                ScopeRow(
                    source_file=str(path),
                    sheet=worksheet.title,
                    row_number=row_number,
                    corpus=detect_corpus(path, worksheet.title, row_text),
                    category=classify(row_text),
                    position=best_position(values),
                    unit=detect_unit(values),
                    quantity=detect_quantity(values, unit_index),
                    formula_or_note="",
                    row_text=row_text,
                )
            )
    return rows


def write_csv(rows: Sequence[ScopeRow], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = list(asdict(rows[0]).keys()) if rows else list(ScopeRow.__annotations__.keys())
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_xlsx(rows: Sequence[ScopeRow], files: Sequence[Path], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    ws_rows = workbook.create_sheet("Rows")
    ws_files = workbook.create_sheet("Files")

    summary = defaultdict(float)
    for row in rows:
        if row.quantity is None:
            continue
        key = (row.corpus, row.category, row.unit)
        summary[key] += row.quantity

    ws_summary.append(["Corpus", "Category", "Unit", "Quantity"])
    for (corpus, category, unit), quantity in sorted(summary.items()):
        ws_summary.append([corpus, category, unit, quantity])

    headers = list(ScopeRow.__annotations__.keys())
    ws_rows.append(headers)
    for row in rows:
        ws_rows.append([getattr(row, header) for header in headers])

    ws_files.append(["Source file", "Size bytes"])
    for path in files:
        ws_files.append([str(path), path.stat().st_size if path.exists() else None])

    for worksheet in (ws_summary, ws_rows, ws_files):
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            max_len = min(70, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2)
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(10, max_len)

    workbook.save(output)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Extract estimate/VOR scope rows from XLSX files.")
    parser.add_argument("--input", required=True, help="Input .xlsx file or folder.")
    parser.add_argument("--output", required=True, help="Output .xlsx or .csv file.")
    parser.add_argument(
        "--keywords",
        default=",".join(DEFAULT_KEYWORDS),
        help="Comma-separated keyword list. Defaults to facade and metal scope terms.",
    )
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    output_path = Path(args.output)
    keywords = [item.strip() for item in args.keywords.split(",") if item.strip()]
    files = list(iter_xlsx_files(input_path))
    all_rows: list[ScopeRow] = []
    for file_path in files:
        all_rows.extend(extract_file(file_path, keywords))

    if output_path.suffix.lower() == ".csv":
        write_csv(all_rows, output_path)
    else:
        write_xlsx(all_rows, files, output_path)

    print(f"Files scanned: {len(files)}")
    print(f"Rows extracted: {len(all_rows)}")
    print(f"Output: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
