import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[3]
SCRIPT = ROOT / "tools" / "estimate_scope" / "extract_estimate_scope.py"


def build_sample(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ВОР 02-01-01"
    sheet.append(["№", "Наименование", "Ед.", "Кол-во", "Формула"])
    sheet.append([37, "Устройство вентилируемых фасадов с облицовкой плитами из керамогранита", "м2", 3737.8, ""])
    sheet.append([43, "Гранитная плита CSG-018 Talia Grey Granite", "м2", 290, "2,9*100"])
    sheet.append([25, "Каркас обрамления проемов с устройством откосов", "м2", 2496, ""])
    workbook.save(path)


def test_extract_sample_workbook() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        source = tmp_path / "sample.xlsx"
        output = tmp_path / "scope.xlsx"
        build_sample(source)

        subprocess.run(
            [sys.executable, str(SCRIPT), "--input", str(source), "--output", str(output)],
            check=True,
            cwd=ROOT,
        )

        workbook = load_workbook(output, data_only=True)
        rows = list(workbook["Rows"].iter_rows(values_only=True))
        summary = list(workbook["Summary"].iter_rows(values_only=True))

        assert len(rows) == 4
        categories = {row[4] for row in rows[1:]}
        assert "NVF / porcelain stoneware" in categories
        assert "Plinth / granite / stone" in categories
        assert "Opening slopes / framing" in categories
        assert any(row[1] == "Plinth / granite / stone" and row[3] == 290 for row in summary[1:])


if __name__ == "__main__":
    test_extract_sample_workbook()
    print("ok")
